"""
crs.views.submission — Submission management.

Functions:
    submission_list            — List view at /crs/submissions/.
    submission_start           — Create-draft form at /crs/submissions/start/.
    submission_detail          — Draft workbench at /crs/submissions/<pk>/.
    submission_save            — POST: save Message Header changes.
    submission_upload_excel    — POST: upload (or replace) the customer data Excel.
    submission_remove_excel    — POST: clear the uploaded Excel.
    submission_excel_download  — GET:  authenticated download of the uploaded Excel.
    submission_delete          — POST: delete a draft submission.
"""
import io
import json
import re
from collections import OrderedDict
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.files.base import ContentFile
from django.db import transaction
from django.http import FileResponse, Http404, HttpResponseNotAllowed, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import load_workbook

from crs.forms import (
    SubmissionExcelUploadForm,
    SubmissionHeaderForm,
    SubmissionStartForm,
)
from crs.models import CountryConfiguration, ReportingFI, Submission, SubmissionXMLFile
from crs.services.parser import parse, ParseResult, Error as ParseError
from crs.services.tokens import resolve_strict
from crs.services import xml_builder, parser
from crs.services.xml_validator import validate as validate_xml


# ===========================================================================
# List + Start
# ===========================================================================
@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_list(request):
    """List all submissions, newest first."""
    submissions = (
        Submission.objects
        .select_related("reporting_fi", "country_config")
        .order_by("-created_at")
    )
    return render(request, "crs/submission_list.html", {"submissions": submissions})


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_start(request):
    """Create a new draft submission."""
    if request.method == "POST":
        form = SubmissionStartForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                fi_obj     = form.cleaned_data["reporting_fi"]
                country    = form.cleaned_data["country_config"]
                year       = form.cleaned_data["year"]
                sending_in = form.cleaned_data["sending_company_in"]
                is_nil     = form.cleaned_data["is_nil_return"]

                ref_id = resolve_strict(
                    country.message_ref_id_template,
                    {"SENDING_FI_IN": sending_in, "YEAR": year},
                )

                submission = Submission.objects.create(
                    reporting_fi=fi_obj,
                    country_config=country,
                    year=year,
                    sending_company_in=sending_in,
                    # TC always = FI's residence country.
                    transmitting_country=fi_obj.res_country_code,
                    # RC default mirrors TC — correct for combined_domestic
                    # strategy. For split_by_residence the stored value is
                    # ignored; XML generator sets RC per file at output time.
                    receiving_country=fi_obj.res_country_code,
                    document_type="OECD1",
                    message_type="CRS703" if is_nil else "CRS701",
                    is_nil_return=is_nil,
                    message_ref_id=ref_id,
                    status="draft",
                    created_by=request.user,
                )
            messages.success(
                request,
                f"Draft submission created. MessageRefID: {ref_id}"
            )
            return redirect("crs:submission_detail", pk=submission.pk)
    else:
        form = SubmissionStartForm()

    fi_tins = {}
    for fi_obj in ReportingFI.objects.filter(is_active=True).prefetch_related("ins"):
        fi_tins[fi_obj.pk] = [
            {"value": ifn.in_value, "issued_by": ifn.issued_by}
            for ifn in fi_obj.ins.all() if ifn.in_type == "TIN"
        ]

    current_year = datetime.now().year
    year_choices = list(range(current_year, current_year - 6, -1))

    return render(request, "crs/submission_start.html", {
        "form": form,
        "fi_tins_json": json.dumps(fi_tins),
        "year_choices": year_choices,
        "default_year": current_year - 1,
        "title": "Start Submission",
    })


# ===========================================================================
# Detail / workbench
# ===========================================================================
def _file_size_safe(file_field):
    """Return a FieldFile's size in bytes, or None if missing from storage."""
    if not file_field:
        return None
    try:
        return file_field.size
    except (OSError, ValueError):
        return None


def _row_to_account_label(accounts):
    """Map (sheet, row) -> a human label '{account#} — {holder}' for the
    account this row belongs to. Both the primary row and all CP rows of an
    account map to the same label, so errors on the parent and any continuation
    rows cluster together under one header."""
    mapping = {}
    for acct in accounts:
        if acct.holder_name:
            who = acct.holder_name
        else:
            who = f"{acct.holder_first_name or ''} {acct.holder_last_name or ''}".strip()
        label = f"{acct.account_number} — {who}" if who else acct.account_number
        mapping[(acct.sheet, acct.row_number)] = label
        for cp in acct.controlling_persons:
            mapping[(acct.sheet, cp.row_number)] = label
    return mapping


def _group_by_account(items, account_label_map, row_context):
    """Group errors/corrections. Successfully parsed rows cluster under their
    account label. Rows that failed to fully parse cluster by (sheet, row) with
    any identifying info the parser captured (Account #, holder/CP name)."""
    groups = OrderedDict()
    for item in items:
        account_label = account_label_map.get((item.sheet, item.row))
        if account_label:
            label = account_label
        elif item.row > 0:
            ctx = row_context.get((item.sheet, item.row), "")
            label = f"{item.sheet} Row {item.row}"
            if ctx:
                label = f"{label} — {ctx}"
        else:
            label = "Workbook"
        groups.setdefault(label, []).append(item)
    return list(groups.items())


def _error_is_fixable(e):
    """True if an error can be corrected with a single-cell value edit.
    Structural problems (holder-type rules, header mismatches, orphan CP
    rows, workbook-level failures) must be fixed in Excel instead."""
    if e.sheet not in ("Individual", "Organisation"):
        return False
    if e.row <= 3:
        return False
    if e.field in ("(header)", "Account Holder Type"):
        return False
    if not e.col or not e.col.isalpha():
        return False
    if e.field == "Account Number" and "no Account Number" in e.reason:
        return False
    return True


def _render_detail(request, submission, header_form=None, excel_form=None):
    """Shared renderer for submission_detail and POST handlers that need to
    re-render with form errors. Re-parses the uploaded Excel on every render
    so validation results always reflect the current file."""
    is_draft = submission.status == "draft"

    parse_result = None
    if submission.uploaded_excel:
        try:
            with submission.uploaded_excel.open("rb") as fh:
                parse_result = parse(fh)
        except FileNotFoundError:
            parse_result = ParseResult()
            parse_result.errors.append(ParseError(
                "(file)", 0, "-", "(storage)", "",
                "uploaded file is missing from storage",
            ))

    # XSD-validate the generated XML on every render so the result is always
    # current. Cheap (<100ms typical) and means sub-step 8's "Close" gate can
    # check validity without orchestrating a separate revalidation step.
    xml_validations = []
    for xml_file in submission.xml_files.all():
        try:
            with xml_file.file.open("rb") as fh:
                result = validate_xml(fh.read())
        except FileNotFoundError:
            continue
        xml_validations.append({
            "receiving_country": xml_file.receiving_country,
            "filename":          xml_file.file.name.rsplit("/", 1)[-1],
            "result":            result,
        })

    any_structural_errors = False
    if parse_result:
        for e in parse_result.errors:
            e.fixable = _error_is_fixable(e)
        any_structural_errors = any(not e.fixable for e in parse_result.errors)

    return render(request, "crs/submission_detail.html", {
        "submission":   submission,
        "header_form":  header_form or SubmissionHeaderForm(instance=submission),
        "excel_form":   excel_form  or SubmissionExcelUploadForm(instance=submission),
        "is_draft":     is_draft,
        "excel_size":   _file_size_safe(submission.uploaded_excel),
        "excel_display_name": (
            submission.uploaded_excel.name.rsplit("/", 1)[-1]
            if submission.uploaded_excel else None
        ),
        "xml_validations":  xml_validations,
        "xml_all_valid":    bool(xml_validations) and all(v["result"].is_valid for v in xml_validations),
        "parse_result":      parse_result,
        "error_groups":      _group_by_account(parse_result.errors, _row_to_account_label(parse_result.accounts), parse_result.row_context) if parse_result else [],
        "correction_groups": _group_by_account(parse_result.corrections, _row_to_account_label(parse_result.accounts), parse_result.row_context) if parse_result else [],
        "any_structural_errors": any_structural_errors,
    })

@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_detail(request, pk):
    """Draft workbench / read-only post-close detail page."""
    submission = get_object_or_404(
        Submission.objects.select_related("reporting_fi", "country_config"),
        pk=pk,
    )
    return _render_detail(request, submission)


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_save(request, pk):
    """Save Message Header edits on a draft submission."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "draft":
        messages.error(
            request,
            f"Only draft submissions can be edited "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    form = SubmissionHeaderForm(request.POST, instance=submission)
    if form.is_valid():
        form.save()
        messages.success(request, "Message Header updated.")
        return redirect("crs:submission_detail", pk=pk)

    # Re-render with errors
    return _render_detail(request, submission, header_form=form)


# ===========================================================================
# Excel upload / replace / remove / download
# ===========================================================================
@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_upload_excel(request, pk):
    """Upload (or replace) the customer-data Excel for a draft submission."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "draft":
        messages.error(
            request,
            f"Only draft submissions can accept Excel uploads "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    # Capture the existing storage path (if any) so we can purge it after
    # the form save points uploaded_excel at the new file.
    old_path = submission.uploaded_excel.name if submission.uploaded_excel else None

    form = SubmissionExcelUploadForm(request.POST, request.FILES, instance=submission)
    if form.is_valid():
        form.save()
        if old_path and old_path != submission.uploaded_excel.name:
            try:
                submission.uploaded_excel.storage.delete(old_path)
            except Exception:
                pass  # File may already be gone; tolerate.
        # Any previously-generated XML is now stale (built from the old Excel).
        # Purge all generated XML files so the user is forced to regenerate
        # from the new data.
        had_files = submission.xml_files.exists()
        for xml_file in submission.xml_files.all():
            try:
                xml_file.file.delete(save=False)
            except Exception:
                pass
            xml_file.delete()
        if had_files:
            submission.xml_generated_at = None
            submission.save(update_fields=["xml_generated_at", "updated_at"])
        filename = submission.uploaded_excel.name.rsplit("/", 1)[-1]
        messages.success(request, f"Excel '{filename}' uploaded.")
    else:
        for error_list in form.errors.values():
            messages.error(request, " ".join(error_list))

    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_remove_excel(request, pk):
    """Clear the uploaded Excel from a draft submission."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "draft":
        messages.error(
            request,
            f"Only draft submissions can have files removed "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    if submission.uploaded_excel:
        try:
            submission.uploaded_excel.delete(save=False)
        except Exception:
            pass
        submission.uploaded_excel = None
        # Drop the generated XML too — it corresponded to the Excel we just
        # removed and is now meaningless.
        update_fields = ["uploaded_excel", "updated_at"]
        if submission.xml_files.exists():
            for xml_file in submission.xml_files.all():
                try:
                    xml_file.file.delete(save=False)
                except Exception:
                    pass
                xml_file.delete()
            submission.xml_generated_at = None
            update_fields += ["xml_generated_at"]
        submission.save(update_fields=update_fields)
        messages.success(request, "Excel file removed.")
    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_excel_download(request, pk):
    """Authenticated download of the uploaded Excel."""
    submission = get_object_or_404(Submission, pk=pk)
    if not submission.uploaded_excel:
        raise Http404("No file uploaded.")
    filename = submission.uploaded_excel.name.rsplit("/", 1)[-1]
    return FileResponse(
        submission.uploaded_excel.open("rb"),
        as_attachment=True,
        filename=filename,
    )


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_fix_cell(request, pk):
    """POST: apply a single-cell value fix to the uploaded workbook.

    Reads the workbook through the storage API (works on local and remote
    storage), writes one cell, saves it back, and purges any now-stale
    generated XML. Re-parse-on-render then reflects the fix. Scoped to
    single-cell value fixes; structural errors are corrected in Excel."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)

    if submission.status != "draft":
        messages.error(request, "Only draft submissions can be edited.")
        return redirect("crs:submission_detail", pk=pk)
    if submission.is_nil_return or not submission.uploaded_excel:
        messages.error(request, "No uploaded workbook to edit.")
        return redirect("crs:submission_detail", pk=pk)

    sheet = (request.POST.get("sheet") or "").strip()
    cell  = (request.POST.get("cell") or "").strip().upper()
    value = (request.POST.get("value") or "").strip()

    if sheet not in ("Individual", "Organisation"):
        messages.error(request, "Invalid sheet reference.")
        return redirect("crs:submission_detail", pk=pk)

    m = re.fullmatch(r"([A-Z]{1,3})([0-9]+)", cell)
    if not m:
        messages.error(request, "Invalid cell reference.")
        return redirect("crs:submission_detail", pk=pk)
    col_letters, row_num = m.group(1), int(m.group(2))
    if row_num <= 3:
        messages.error(request, "Header rows cannot be edited here.")
        return redirect("crs:submission_detail", pk=pk)

    # Read -> modify -> write back through the storage API (storage-agnostic).
    try:
        with submission.uploaded_excel.open("rb") as fh:
            data = fh.read()
        wb = load_workbook(io.BytesIO(data))
        if sheet not in wb.sheetnames:
            messages.error(request, f"Sheet '{sheet}' not found in the workbook.")
            return redirect("crs:submission_detail", pk=pk)
        ws = wb[sheet]
        target = ws[f"{col_letters}{row_num}"]
        target.value = value or None
        # Force text so a fixed code/date isn't re-coerced by Excel on open.
        target.number_format = "@"
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
    except Exception as exc:
        messages.error(request, f"Could not apply fix: {exc}")
        return redirect("crs:submission_detail", pk=pk)

    old_name = submission.uploaded_excel.name
    base_name = old_name.rsplit("/", 1)[-1]
    submission.uploaded_excel.save(base_name, ContentFile(out.read()), save=False)
    submission.save(update_fields=["uploaded_excel", "updated_at"])
    if old_name and old_name != submission.uploaded_excel.name:
        try:
            submission.uploaded_excel.storage.delete(old_name)
        except Exception:
            pass

    # A cell change invalidates any previously-generated XML — purge it.
    if submission.xml_files.exists():
        for xf in submission.xml_files.all():
            try:
                xf.file.delete(save=False)
            except Exception:
                pass
            xf.delete()
        submission.xml_generated_at = None
        submission.save(update_fields=["xml_generated_at", "updated_at"])

    messages.success(
        request,
        f"{sheet} {col_letters}{row_num} updated"
        + (f" to '{value}'." if value else " (cleared).")
    )
    return redirect("crs:submission_detail", pk=pk)


# ===========================================================================
# Delete
# ===========================================================================
@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_delete(request, pk):
    """Delete a submission of any status. POST-only.

    Drafts: cleans up uploaded Excel and any generated XML files from storage.
    Post-draft states: the close lifecycle has already purged Excel + XML,
    but we still iterate xml_files defensively in case storage drift left
    orphan blobs. The DB row (and any FK'd xml_files records) cascade-delete."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission   = get_object_or_404(Submission, pk=pk)
    ref_id       = submission.message_ref_id
    status_label = submission.get_status_display()

    # Clean up uploaded Excel from storage if any (drafts)
    if submission.uploaded_excel:
        try:
            submission.uploaded_excel.delete(save=False)
        except Exception:
            pass

    # Clean up any generated XML files from storage (drafts; defensive for others)
    for xml_file in submission.xml_files.all():
        if xml_file.file:
            try:
                xml_file.file.delete(save=False)
            except Exception:
                pass

    submission.delete()
    messages.success(
        request,
        f"Submission '{ref_id}' ({status_label}) deleted."
    )
    return redirect("crs:submission_list")


# ===========================================================================
# XML generation + download
# ===========================================================================
@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_generate_xml(request, pk):
    """POST: (re)generate all XML files for this submission.

    For combined_domestic: produces exactly one file with RC = FI residence.
    For split_by_residence: produces N files, one per reportable RC.

    Always wipes any pre-existing files for the submission first to guarantee
    a clean re-generation."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)

    if submission.status != "draft":
        messages.error(request, "Cannot regenerate XML — submission is not in draft state.")
        return redirect("crs:submission_detail", pk=pk)

    # Nil return: no Excel, no parse — build a single CRS703 file directly.
    if submission.is_nil_return:
        try:
            files_by_rc = xml_builder.build_nil(submission)
        except Exception as exc:
            messages.error(request, f"XML build failed: {exc}")
            return redirect("crs:submission_detail", pk=pk)
    else:
        if not submission.uploaded_excel:
            messages.error(request, "Upload an Excel file first.")
            return redirect("crs:submission_detail", pk=pk)

        # Re-parse so we always work from the workbook's current state
        try:
            parse_result = parser.parse(submission.uploaded_excel.path)
        except Exception as exc:
            messages.error(request, f"Parse failed: {exc}")
            return redirect("crs:submission_detail", pk=pk)

        if not parse_result.is_valid:
            messages.error(
                request,
                f"Cannot generate XML — Excel parser reported "
                f"{len(parse_result.errors)} error(s). Fix the workbook and retry."
            )
            return redirect("crs:submission_detail", pk=pk)

        try:
            files_by_rc = xml_builder.build_all(submission, parse_result)
        except Exception as exc:
            messages.error(request, f"XML build failed: {exc}")
            return redirect("crs:submission_detail", pk=pk)

        if not files_by_rc:
            messages.error(
                request,
                "No reportable accounts in the workbook — nothing to generate. "
                "Check that holders aren't all US or domestic residents."
            )
            return redirect("crs:submission_detail", pk=pk)

    # Filename template handling. For split mode, ensure [RECEIVING_COUNTRY]
    # appears in the template so split files don't collide. Auto-prepend if
    # the user hasn't added the token to their config.
    cc = submission.country_config
    base_template = cc.output_filename_template
    if cc.receiving_country_strategy == "split_by_residence" \
            and "[RECEIVING_COUNTRY]" not in base_template:
        base_template = "[RECEIVING_COUNTRY]_" + base_template

    # Wipe any existing files for this submission first
    for existing in submission.xml_files.all():
        try:
            existing.file.delete(save=False)
        except Exception:
            pass
        existing.delete()

    # Save fresh files
    from django.core.files.base import ContentFile
    for rc, xml_bytes in files_by_rc.items():
        filename = resolve_strict(base_template, {
            "SENDING_FI_IN":     submission.sending_company_in,
            "YEAR":              submission.year,
            "RECEIVING_COUNTRY": rc,
        })
        # Count AccountReport elements as cheap metadata for the UI
        account_count = xml_bytes.count(b"<crs:AccountReport>")

        xml_file = SubmissionXMLFile(
            submission=submission,
            receiving_country=rc,
            size=len(xml_bytes),
            account_count=account_count,
        )
        xml_file.file.save(filename, ContentFile(xml_bytes), save=False)
        xml_file.save()

    submission.xml_generated_at = timezone.now()
    submission.save(update_fields=["xml_generated_at", "updated_at"])

    n = len(files_by_rc)
    rc_list = ", ".join(sorted(files_by_rc.keys()))
    messages.success(
        request,
        f"Generated {n} XML file{'s' if n != 1 else ''} ({rc_list})."
    )
    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_xml_file_download(request, pk, xml_file_id):
    """Authenticated download of one specific XML file by primary key.
    Validates the file belongs to the supplied submission to prevent
    cross-submission leakage via URL guessing."""
    submission = get_object_or_404(Submission, pk=pk)
    xml_file = get_object_or_404(
        SubmissionXMLFile, pk=xml_file_id, submission=submission,
    )
    filename = xml_file.file.name.rsplit("/", 1)[-1]
    return FileResponse(
        xml_file.file.open("rb"),
        as_attachment=True,
        filename=filename,
        content_type="application/xml",
    )


@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_xml_file_view(request, pk, xml_file_id):
    """Return the XML file's content inline (not as attachment) so it can
    be fetched via JS and displayed in the modal viewer without forcing
    a download. Same auth gates as the download view."""
    submission = get_object_or_404(Submission, pk=pk)
    xml_file = get_object_or_404(
        SubmissionXMLFile, pk=xml_file_id, submission=submission,
    )
    try:
        with xml_file.file.open("rb") as fh:
            content = fh.read()
    except FileNotFoundError:
        return HttpResponse("File not found in storage.", status=404)
    return HttpResponse(content, content_type="application/xml; charset=utf-8")


@login_required
@permission_required('auth.can_access_crs', raise_exception=True)
def submission_xml_files_download_all(request, pk):
    """Stream all XML files for a submission as a single ZIP archive.
    Useful for split-by-residence submissions that fan out to many files —
    one click instead of N. Missing files (storage drift) are silently
    skipped rather than aborting the whole archive."""
    import io
    import zipfile

    submission = get_object_or_404(Submission, pk=pk)
    xml_files = list(submission.xml_files.all())
    if not xml_files:
        messages.error(request, "No XML files to download.")
        return redirect("crs:submission_detail", pk=pk)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for xml_file in xml_files:
            filename = xml_file.file.name.rsplit("/", 1)[-1]
            try:
                with xml_file.file.open("rb") as fh:
                    zf.writestr(filename, fh.read())
            except FileNotFoundError:
                continue
    buf.seek(0)

    zip_name = f"{submission.sending_company_in}_{submission.year}_xml_files.zip"
    return FileResponse(
        buf,
        as_attachment=True,
        filename=zip_name,
        content_type="application/zip",
    )


# ===========================================================================
# Lifecycle transitions
# ===========================================================================
def _append_note(submission, label, text):
    """Append a labelled note to acknowledgement_notes. Skips silently if
    text is empty. Format: '[Label] free text', blank-line separated."""
    text = (text or "").strip()
    if not text:
        return
    entry = f"[{label}] {text}"
    if submission.acknowledgement_notes:
        submission.acknowledgement_notes = (
            submission.acknowledgement_notes + "\n\n" + entry
        )
    else:
        submission.acknowledgement_notes = entry


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_close(request, pk):
    """Close a draft: validate the generated XML one last time, purge the
    transient uploaded Excel, freeze the XML, transition to 'closed'.
    Gated on XSD validity — the whole point of closing is to lock in a
    submittable file, so we re-validate at close time rather than trusting
    the cached banner state in the UI."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)

    if submission.status != "draft":
        messages.error(
            request,
            f"Only draft submissions can be closed "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    # Re-validate every generated file at close time — don't trust the
    # cached banner. Any single file failing blocks the close.
    xml_files = list(submission.xml_files.all())
    if not xml_files:
        messages.error(request, "No generated XML to close. Generate XML first.")
        return redirect("crs:submission_detail", pk=pk)

    for xml_file in xml_files:
        try:
            with xml_file.file.open("rb") as fh:
                validation = validate_xml(fh.read())
        except FileNotFoundError:
            messages.error(
                request,
                f"Generated XML for {xml_file.receiving_country} is missing "
                f"from storage. Regenerate first."
            )
            return redirect("crs:submission_detail", pk=pk)

        if not validation.is_valid:
            messages.error(
                request,
                f"Cannot close — XML for {xml_file.receiving_country} failed "
                f"XSD validation ({len(validation.errors)} error"
                f"{'s' if len(validation.errors) != 1 else ''}). Regenerate "
                f"from a clean Excel and try again."
            )
            return redirect("crs:submission_detail", pk=pk)

    # Purge BOTH the transient Excel AND the generated XML at close. Both
    # contain PII (account holders, TINs, balances, CPs) and the user has
    # the XML they need locally by this point. Submission metadata, lifecycle
    # timestamps, and notes persist — the data payload does not.
    if submission.uploaded_excel:
        try:
            submission.uploaded_excel.delete(save=False)
        except Exception:
            pass
    submission.uploaded_excel = None

    # Purge all generated XML files (one for combined, N for split). Each
    # SubmissionXMLFile carries its own storage file; delete both the file
    # and the row.
    for xml_file in submission.xml_files.all():
        try:
            xml_file.file.delete(save=False)
        except Exception:
            pass
        xml_file.delete()

    submission.status = "closed"
    submission.closed_at = timezone.now()
    submission.save(update_fields=[
        "status", "closed_at", "uploaded_excel", "updated_at",
    ])

    messages.success(
        request,
        "Submission closed. Excel and XML both purged; submission record retains "
        "metadata and lifecycle history only."
    )
    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_mark_submitted(request, pk):
    """Stamp a closed submission as filed with the tax authority externally.
    Optional free-text notes (reference number, portal receipt, etc.)
    appended to acknowledgement_notes with a [Submitted] tag."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "closed":
        messages.error(
            request,
            f"Can only mark as submitted from Closed state "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    submission.status = "submitted_externally"
    submission.submitted_externally_at = timezone.now()
    _append_note(submission, "Submitted", request.POST.get("notes"))
    submission.save(update_fields=[
        "status", "submitted_externally_at", "acknowledgement_notes", "updated_at",
    ])
    messages.success(request, "Marked as submitted externally.")
    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_mark_acknowledged(request, pk):
    """Stamp a submitted_externally submission as acknowledged by the
    tax authority. Optional notes appended with [Acknowledged] tag.
    Terminal state — no further transitions."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "submitted_externally":
        messages.error(
            request,
            f"Can only mark as acknowledged from Submitted Externally state "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    submission.status = "acknowledged"
    submission.acknowledged_at = timezone.now()
    _append_note(submission, "Acknowledged", request.POST.get("notes"))
    submission.save(update_fields=[
        "status", "acknowledged_at", "acknowledgement_notes", "updated_at",
    ])
    messages.success(request, "Marked as acknowledged.")
    return redirect("crs:submission_detail", pk=pk)


@login_required
@permission_required('auth.can_edit_crs', raise_exception=True)
def submission_mark_rejected(request, pk):
    """Stamp a submitted_externally submission as rejected by the tax
    authority. Rejection reason notes are REQUIRED (we want a record of
    why). Terminal state — no further transitions; user creates a fresh
    draft to retry."""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    submission = get_object_or_404(Submission, pk=pk)
    if submission.status != "submitted_externally":
        messages.error(
            request,
            f"Can only mark as rejected from Submitted Externally state "
            f"(current status: {submission.get_status_display()})."
        )
        return redirect("crs:submission_detail", pk=pk)

    reason = (request.POST.get("notes") or "").strip()
    if not reason:
        messages.error(request, "A rejection reason is required.")
        return redirect("crs:submission_detail", pk=pk)

    submission.status = "rejected"
    submission.acknowledged_at = timezone.now()
    _append_note(submission, "Rejected", reason)
    submission.save(update_fields=[
        "status", "acknowledged_at", "acknowledgement_notes", "updated_at",
    ])
    messages.success(request, "Marked as rejected.")
    return redirect("crs:submission_detail", pk=pk)
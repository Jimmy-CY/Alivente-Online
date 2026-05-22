"""
crs.services.parser — Parse and validate the uploaded customer-data Excel.

Public surface:
    parse(file_obj) -> ParseResult

ParseResult bundles three lists:
    .accounts    — successfully parsed ParsedAccount objects
    .corrections — auto-fixes applied silently (sheet, row, col, field, original, corrected, reason)
    .errors      — blocking validation failures (sheet, row, col, field, value, reason)
    .is_valid    — True iff errors is empty
    .row_context — (sheet, row) -> short identifying label for UI group headers

A submission can only proceed to XML generation when .is_valid.

Sheet conventions:
    - Row 1: sheet title (skipped)
    - Row 2: section headers (skipped)
    - Row 3: column headers (validated against expected layout)
    - Row 4+: data
    - Column A: visual marker only (the 'sample →' on Individual,
      a CRS101/102/103 indicator on Organisation). Real data starts at column B.

Organisation multi-row Controlling Person pattern:
    A 'new account' row has Account Number (col B) filled AND account-level
    columns (C-X) populated. A 'continuation row' repeats the same Account
    Number in B, leaves C-X blank, and carries one additional Controlling
    Person in Y-AG. Multiple continuation rows can attach to one account.

Account Holder Type — Controlling Person rules (validated post-parse):
    CRS101 — Passive NFE with Reportable CPs:  MUST have >= 1 CP.
    CRS102 — Reportable Entity (not a Passive NFE):  must NOT have CPs.
             If you have reportable CPs as well, use CRS103 instead.
    CRS103 — Passive NFE that is itself Reportable:  CPs allowed, optional.

    These rules are evaluated against an independent per-row observation
    tracker (not against result.accounts), so they fire even when the same
    row has other field-level errors that prevented a ParsedAccount being
    produced. CP count reflects intent (cells populated in Y..AG), not
    successful CP parse, so a CRS102 row with a broken CP still surfaces
    the holder-type mistake instead of hiding it behind the CP errors.
"""
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Optional

from openpyxl import load_workbook

from crs.services import reference_data as ref
from crs.services.normalize import (
    normalize_amount, normalize_boolean, normalize_choice,
    normalize_country, normalize_currency, normalize_date, normalize_string,
)


# ---------------------------------------------------------------------------
# Expected column headers (row 3) — used to validate sheet structure
# ---------------------------------------------------------------------------
INDIVIDUAL_HEADERS = [
    None,                          # A — marker column
    "Account Number",              # B
    "Account Type",                # C
    "Closed",                      # D
    "Dormant",                     # E
    "Undocumented",                # F
    "First Name",                  # G
    "Last Name",                   # H
    "Tax Residence",               # I
    "TIN",                         # J
    "Issued By",                   # K
    "Birth Date (yyyy-mm-dd)",     # L
    "Country Code",                # M
    "Free Address",                # N
    "Account Balance",             # O
    "Currency",                    # P
    "Dividend",       "Currency",  # Q, R
    "Interest",       "Currency",  # S, T
    "Gross Proceeds", "Currency",  # U, V
    "Other Payment",  "Currency",  # W, X
]

ORGANISATION_HEADERS = [
    None,                          # A — marker column
    "Account Number",              # B
    "Account Type",                # C
    "Closed", "Dormant", "Undocumented",   # D, E, F
    "Name",                        # G
    "Account Holder Type",         # H
    "Tax Residence",               # I
    "IN", "Issued By", "IN Type",  # J, K, L
    "Country Code",                # M
    "Free Address",                # N
    "Account Balance", "Currency", # O, P
    "Dividend",       "Currency",  # Q, R
    "Interest",       "Currency",  # S, T
    "Gross Proceeds", "Currency",  # U, V
    "Other Payment",  "Currency",  # W, X
    "First Name", "Last Name",     # Y, Z
    "Ctrl. Person Type",           # AA
    "Tax Residence",               # AB
    "TIN", "Issued By",            # AC, AD
    "Birth Date (yyyy-mm-dd)",     # AE
    "Country Code",                # AF
    "Free Address",                # AG
]


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class ParsedPayment:
    type: str             # 'Dividend' / 'Interest' / 'GrossProceeds' / 'Other'
    amount: Decimal
    currency: str


@dataclass
class ParsedControllingPerson:
    row_number: int
    first_name: str
    last_name: str
    cp_type: str          # CRS801 .. CRS813
    tax_residence: str
    tin: str
    tin_issued_by: str
    birth_date: str       # ISO YYYY-MM-DD
    address_country: str
    address_free: str


@dataclass
class ParsedAccount:
    sheet: str            # 'Individual' or 'Organisation'
    row_number: int
    account_number: str
    account_number_type: str
    is_closed: bool
    is_dormant: bool
    is_undocumented: bool

    # Holder identity — Individual fields
    holder_first_name: Optional[str] = None
    holder_last_name: Optional[str] = None
    holder_birth_date: Optional[str] = None

    # Holder identity — Organisation fields
    holder_name: Optional[str] = None
    holder_type: Optional[str] = None       # CRS101 / CRS102 / CRS103
    in_type: Optional[str] = None           # TIN / GIIN / EIN / BRN / LEI / Other

    # Common
    tax_residence: str = ""
    tin: str = ""
    tin_issued_by: str = ""
    address_country: str = ""
    address_free: str = ""

    balance: Optional[Decimal] = None
    balance_currency: str = ""

    payments: list = field(default_factory=list)
    controlling_persons: list = field(default_factory=list)


@dataclass
class Correction:
    sheet: str
    row: int
    col: str
    field: str
    original: str
    corrected: str
    reason: str


@dataclass
class Error:
    sheet: str
    row: int
    col: str
    field: str
    value: str
    reason: str


@dataclass
class ParseResult:
    accounts: list = field(default_factory=list)
    corrections: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    # (sheet, row) -> short label like "CY1234... — Maria Demetriou", captured
    # for every non-empty data row regardless of parse success. Used by the UI
    # to enrich error/correction group headers with the row's identity.
    row_context: dict = field(default_factory=dict)

    @property
    def is_valid(self):
        return not self.errors

    @property
    def has_corrections(self):
        return bool(self.corrections)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _col_letter(idx):
    """Convert 1-based column index to letter(s): 1->A, 26->Z, 27->AA, 33->AG."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(ord("A") + r) + s
    return s


def _is_blank(v):
    return v is None or (isinstance(v, str) and not v.strip())


def _row_is_empty(row, start_col=2, end_col=None):
    end = end_col if end_col is not None else len(row)
    return all(_is_blank(row[i]) for i in range(start_col - 1, end))


def _account_section_empty(row):
    """Organisation: True if account-level columns C..X (idx 2..23) are blank."""
    return all(_is_blank(row[i]) for i in range(2, 24))


def _has_cp_data(row):
    """Organisation: True if any CP column Y..AG (idx 24..32) has data."""
    return any(not _is_blank(row[i]) for i in range(24, 33))


def _disp(v):
    """Render a cell value safely as a short display string for error/correction logs."""
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= 80 else s[:77] + "..."


def _safe_str(v, maxlen=40):
    if v is None:
        return ""
    s = str(v).strip()
    return s if len(s) <= maxlen else s[:maxlen - 3] + "..."


def _row_label(account_number, *holder_parts):
    """Build '{account} — {holder}' from raw cell values for use as a group label.
    Either side can be missing; if both are blank, returns ''."""
    acct = _safe_str(account_number)
    holder = " ".join(_safe_str(p) for p in holder_parts if not _is_blank(p)).strip()
    parts = [p for p in (acct, holder) if p]
    return " — ".join(parts)


def _make_extractor(result, sheet, row_num):
    """Return a closure that extracts one cell, applies a normalizer, and records
    any error/correction onto `result` with full sheet/row/column context."""
    def extract(col_idx, field_name, raw, normalizer, required=True, **kwargs):
        value, error, correction = normalizer(raw, **kwargs)
        col = _col_letter(col_idx)
        if error:
            result.errors.append(Error(sheet, row_num, col, field_name, _disp(raw), error))
            return None
        if required and value is None:
            result.errors.append(Error(sheet, row_num, col, field_name, "", "required field is missing"))
            return None
        if correction:
            result.corrections.append(Correction(
                sheet, row_num, col, field_name,
                _disp(raw), _disp(value), correction,
            ))
        return value
    return extract


# ---------------------------------------------------------------------------
# Header validation
# ---------------------------------------------------------------------------
def _validate_headers(ws, sheet_name, expected, result):
    """Check row 3 cells against expected headers. Mismatches are errors but
    we still attempt to parse — the user might have only renamed a header."""
    rows = ws.iter_rows(min_row=3, max_row=3, max_col=len(expected), values_only=True)
    try:
        row_values = next(iter(rows))
    except StopIteration:
        row_values = ()
    for ci, want in enumerate(expected, 1):
        if want is None:
            continue
        got = row_values[ci - 1] if ci - 1 < len(row_values) else None
        got_str = (got or "").strip() if isinstance(got, str) else (str(got) if got is not None else "")
        if got_str != want:
            result.errors.append(Error(
                sheet_name, 3, _col_letter(ci), "(header)", got_str,
                f"expected header '{want}', got '{got_str}'",
            ))


# ---------------------------------------------------------------------------
# Payment pair extraction (shared between Individual and Organisation)
# ---------------------------------------------------------------------------
def _extract_payments(extract, row, start_col):
    """Extract the four optional payment pairs starting at start_col (=17, column Q).
    Each pair is (amount, currency). Both-blank → skip. One-blank → error on the missing."""
    payments = []
    specs = [
        ("Dividend",      start_col,     start_col + 1),
        ("Interest",      start_col + 2, start_col + 3),
        ("GrossProceeds", start_col + 4, start_col + 5),
        ("Other",         start_col + 6, start_col + 7),
    ]
    for ptype, amt_col, curr_col in specs:
        raw_amt = row[amt_col - 1]
        raw_curr = row[curr_col - 1]
        if _is_blank(raw_amt) and _is_blank(raw_curr):
            continue  # neither present → not reported, fine
        amt = extract(amt_col, f"{ptype} Amount", raw_amt, normalize_amount, required=True)
        curr = extract(curr_col, f"{ptype} Currency", raw_curr, normalize_currency, required=True)
        if amt is not None and curr is not None:
            payments.append(ParsedPayment(type=ptype, amount=amt, currency=curr))
    return payments


# ---------------------------------------------------------------------------
# Individual sheet
# ---------------------------------------------------------------------------
def _parse_individual_sheet(ws, result):
    for ri, row_tuple in enumerate(
        ws.iter_rows(min_row=4, max_col=24, values_only=True), start=4
    ):
        row = list(row_tuple)
        if _row_is_empty(row, start_col=2):
            continue
        # Skip the 'sample →' annotation row
        marker = row[0]
        if isinstance(marker, str) and marker.strip().lower().startswith("sample"):
            continue

        # Capture identifying info (Account #, First + Last) for the group label
        result.row_context[("Individual", ri)] = _row_label(row[1], row[6], row[7])

        account = _parse_individual_row(row, ri, result)
        if account is not None:
            result.accounts.append(account)


def _parse_individual_row(row, row_num, result):
    e = _make_extractor(result, "Individual", row_num)

    acct_num    = e( 2, "Account Number",     row[ 1], normalize_string, max_length=200)
    acct_type   = e( 3, "Account Type",       row[ 2], normalize_choice,
                     valid_set=ref.ACCOUNT_TYPES, label="Account Type")
    is_closed   = e( 4, "Closed",             row[ 3], normalize_boolean)
    is_dormant  = e( 5, "Dormant",            row[ 4], normalize_boolean)
    is_undoc    = e( 6, "Undocumented",       row[ 5], normalize_boolean)
    first_name  = e( 7, "First Name",         row[ 6], normalize_string)
    last_name   = e( 8, "Last Name",          row[ 7], normalize_string)
    tax_res     = e( 9, "Tax Residence",      row[ 8], normalize_country)
    tin         = e(10, "TIN",                row[ 9], normalize_string)
    tin_issuer  = e(11, "TIN Issued By",      row[10], normalize_country, required=False)
    birth       = e(12, "Birth Date",         row[11], normalize_date)
    addr_cc     = e(13, "Address Country",    row[12], normalize_country)
    addr_free   = e(14, "Free Address",       row[13], normalize_string, max_length=4000)
    balance     = e(15, "Account Balance",    row[14], normalize_amount)
    bal_curr    = e(16, "Balance Currency",   row[15], normalize_currency)

    payments = _extract_payments(e, row, start_col=17)

    # Bail if any required field failed (errors already logged by extractor)
    required = [acct_num, acct_type, is_closed, is_dormant, is_undoc,
                first_name, last_name, tax_res, tin, birth,
                addr_cc, addr_free, balance, bal_curr]
    if any(v is None for v in required):
        return None

    return ParsedAccount(
        sheet="Individual", row_number=row_num,
        account_number=acct_num, account_number_type=acct_type,
        is_closed=is_closed, is_dormant=is_dormant, is_undocumented=is_undoc,
        holder_first_name=first_name, holder_last_name=last_name,
        holder_birth_date=birth,
        tax_residence=tax_res, tin=tin, tin_issued_by=tin_issuer or "",
        address_country=addr_cc, address_free=addr_free,
        balance=balance, balance_currency=bal_curr,
        payments=payments,
    )


# ---------------------------------------------------------------------------
# Organisation sheet
# ---------------------------------------------------------------------------
def _parse_organisation_sheet(ws, result):
    # In-progress ParsedAccount; becomes None when the last primary row's
    # required-field parse failed. Field-level errors are still logged by the
    # extractor, but no ParsedAccount makes it into result.accounts for that row.
    current = None

    # Independent observation tracker. Captures (row, acct_num, holder_type,
    # cp_count) for every primary Organisation row processed, regardless of
    # whether the row produces a ParsedAccount. The post-pass CRS101/CRS102
    # holder-type rules iterate this list instead of result.accounts, so they
    # fire even when the same row has other field-level errors.
    #
    # cp_count reflects INTENT (cells populated in Y..AG), not successful CP
    # parse. A CRS102 row with a broken CP should still surface the holder-type
    # mistake — otherwise the user fixes the CP fields only to discover the
    # deeper "wrong holder type" issue on the next round-trip.
    #
    # Each observation: {"row": int, "acct_num": str,
    #                    "holder_type": str|None, "cp_count": int}.
    org_observations = []
    current_obs = None

    for ri, row_tuple in enumerate(
        ws.iter_rows(min_row=4, max_col=33, values_only=True), start=4
    ):
        row = list(row_tuple)
        if _row_is_empty(row, start_col=2, end_col=33):
            continue

        # Capture identifying info for the group label BEFORE any branching,
        # so every non-empty row gets the most informative label we can build
        # regardless of which code path it takes:
        #   - primary row → prefer Org Name (col G, idx 6)
        #   - continuation row → prefer CP First+Last (cols Y, Z, idx 24, 25)
        #   - orphan / partial → fall back to Account Number alone
        if not _is_blank(row[6]):
            result.row_context[("Organisation", ri)] = _row_label(row[1], row[6])
        elif not _is_blank(row[24]) or not _is_blank(row[25]):
            result.row_context[("Organisation", ri)] = _row_label(row[1], row[24], row[25])
        else:
            result.row_context[("Organisation", ri)] = _row_label(row[1])

        acct_num_raw = row[1]  # column B
        if _is_blank(acct_num_raw):
            # CP data without an Account Number = orphan row
            if _has_cp_data(row):
                result.errors.append(Error(
                    "Organisation", ri, "B", "Account Number", "",
                    "row has Controlling Person data but no Account Number",
                ))
            continue

        acct_num_str = str(acct_num_raw).strip()

        # Continuation detection runs off current_obs (not current), so a
        # continuation row is still recognized as such even when its primary
        # row's parse failed. Without this, a continuation following a failed
        # primary would be treated as its own primary — cascading account-level
        # "required field missing" errors AND losing the CP from the count.
        is_continuation = (
            current_obs is not None
            and acct_num_str == current_obs["acct_num"]
            and _account_section_empty(row)
        )

        if is_continuation:
            if _has_cp_data(row):
                current_obs["cp_count"] += 1
                cp = _parse_cp_row(row, ri, result)
                # Attach only if both the CP parsed AND the primary survived.
                # If either failed, the cp_count bump above is still enough
                # for the CRS101/102 holder-type rule to fire correctly.
                if cp is not None and current is not None:
                    current.controlling_persons.append(cp)
            continue

        # New primary row: flush previous account + observation, start fresh.
        if current is not None:
            result.accounts.append(current)
        if current_obs is not None:
            org_observations.append(current_obs)

        # Build the observation BEFORE the primary parse attempt so it's
        # captured even when the parse fails. holder_type is taken raw (just
        # uppercased/stripped) — the normalize_choice call inside the primary
        # parser logs its own error if the value is invalid; the post-pass
        # rule check ignores anything outside the three known codes so we
        # don't pile a misleading "needs CPs" message on top of that.
        holder_type_raw = row[7]  # column H
        if _is_blank(holder_type_raw):
            holder_type_obs = None
        else:
            holder_type_obs = str(holder_type_raw).strip().upper()
        current_obs = {
            "row": ri,
            "acct_num": acct_num_str,
            "holder_type": holder_type_obs,
            # Primary row may carry the first CP inline (cols Y..AG).
            # Count intent (cells populated), not successful parse.
            "cp_count": 1 if _has_cp_data(row) else 0,
        }

        current = _parse_organisation_primary_row(row, ri, result)
        if current is not None and _has_cp_data(row):
            cp = _parse_cp_row(row, ri, result)
            if cp is not None:
                current.controlling_persons.append(cp)

    # End-of-loop flush for both trackers.
    if current is not None:
        result.accounts.append(current)
    if current_obs is not None:
        org_observations.append(current_obs)

    # Account Holder Type — Controlling Person rules per OECD CRS spec:
    #   CRS101 — Passive NFE with Reportable CPs:  MUST have >= 1 CP.
    #   CRS102 — Reportable Entity (not a Passive NFE):  must NOT have CPs.
    #            If you have reportable CPs as well, use CRS103 instead.
    #   CRS103 — Passive NFE that is itself Reportable:  CPs optional.
    for obs in org_observations:
        if obs["holder_type"] not in ("CRS101", "CRS102", "CRS103"):
            # Blank holder_type   → already errored by the required-field check.
            # Invalid (e.g. 'CRS199') → already errored by normalize_choice.
            # Either way, don't cascade a misleading holder-type CP rule on
            # top of the underlying holder-type error.
            continue
        n_cps = obs["cp_count"]
        if obs["holder_type"] == "CRS101" and n_cps == 0:
            result.errors.append(Error(
                "Organisation", obs["row"], "H", "Account Holder Type", "CRS101",
                "CRS101 (Passive NFE with Reportable Controlling Persons) "
                "requires at least one Controlling Person",
            ))
        elif obs["holder_type"] == "CRS102" and n_cps > 0:
            result.errors.append(Error(
                "Organisation", obs["row"], "H", "Account Holder Type", "CRS102",
                f"CRS102 (Reportable Entity, not a Passive NFE) should not have "
                f"Controlling Persons (found {n_cps}). If this entity is also a "
                f"Passive NFE with reportable Controlling Persons, use CRS103 instead.",
            ))
        # CRS103: CPs allowed, optional — no constraint either way.


def _parse_organisation_primary_row(row, row_num, result):
    e = _make_extractor(result, "Organisation", row_num)

    acct_num    = e( 2, "Account Number",     row[ 1], normalize_string, max_length=200)
    acct_type   = e( 3, "Account Type",       row[ 2], normalize_choice,
                     valid_set=ref.ACCOUNT_TYPES, label="Account Type")
    is_closed   = e( 4, "Closed",             row[ 3], normalize_boolean)
    is_dormant  = e( 5, "Dormant",            row[ 4], normalize_boolean)
    is_undoc    = e( 6, "Undocumented",       row[ 5], normalize_boolean)
    name        = e( 7, "Name",               row[ 6], normalize_string)
    holder_type = e( 8, "Account Holder Type", row[ 7], normalize_choice,
                     valid_set=ref.ACCT_HOLDER_TYPES, label="Account Holder Type")
    tax_res     = e( 9, "Tax Residence",      row[ 8], normalize_country)
    in_value    = e(10, "IN",                 row[ 9], normalize_string)
    in_issuer   = e(11, "IN Issued By",       row[10], normalize_country, required=False)
    in_type     = e(12, "IN Type",            row[11], normalize_choice,
                     valid_set=ref.IN_TYPES, label="IN Type", required=False)
    addr_cc     = e(13, "Address Country",    row[12], normalize_country)
    addr_free   = e(14, "Free Address",       row[13], normalize_string, max_length=4000)
    balance     = e(15, "Account Balance",    row[14], normalize_amount)
    bal_curr    = e(16, "Balance Currency",   row[15], normalize_currency)

    payments = _extract_payments(e, row, start_col=17)

    required = [acct_num, acct_type, is_closed, is_dormant, is_undoc,
                name, holder_type, tax_res, in_value,
                addr_cc, addr_free, balance, bal_curr]
    if any(v is None for v in required):
        return None

    return ParsedAccount(
        sheet="Organisation", row_number=row_num,
        account_number=acct_num, account_number_type=acct_type,
        is_closed=is_closed, is_dormant=is_dormant, is_undocumented=is_undoc,
        holder_name=name, holder_type=holder_type,
        tax_residence=tax_res, tin=in_value, tin_issued_by=in_issuer or "",
        in_type=in_type,
        address_country=addr_cc, address_free=addr_free,
        balance=balance, balance_currency=bal_curr,
        payments=payments,
    )


def _parse_cp_row(row, row_num, result):
    e = _make_extractor(result, "Organisation", row_num)

    first_name  = e(25, "CP First Name",        row[24], normalize_string)
    last_name   = e(26, "CP Last Name",         row[25], normalize_string)
    cp_type     = e(27, "Ctrl. Person Type",    row[26], normalize_choice,
                    valid_set=ref.CTRL_PERSON_TYPES, label="Ctrl. Person Type")
    tax_res     = e(28, "CP Tax Residence",     row[27], normalize_country)
    tin         = e(29, "CP TIN",               row[28], normalize_string)
    tin_issuer  = e(30, "CP TIN Issued By",     row[29], normalize_country, required=False)
    birth       = e(31, "CP Birth Date",        row[30], normalize_date)
    addr_cc     = e(32, "CP Address Country",   row[31], normalize_country)
    addr_free   = e(33, "CP Free Address",      row[32], normalize_string, max_length=4000)

    required = [first_name, last_name, cp_type, tax_res, tin, birth, addr_cc, addr_free]
    if any(v is None for v in required):
        return None

    return ParsedControllingPerson(
        row_number=row_num,
        first_name=first_name, last_name=last_name, cp_type=cp_type,
        tax_residence=tax_res, tin=tin, tin_issued_by=tin_issuer or "",
        birth_date=birth,
        address_country=addr_cc, address_free=addr_free,
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def parse(file_obj):
    """Parse the uploaded Excel and return a ParseResult.

    `file_obj` can be a path, a file-like object, or a Django FieldFile
    (anything openpyxl.load_workbook accepts)."""
    result = ParseResult()

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as ex:
        result.errors.append(Error(
            "(workbook)", 0, "-", "(file)", "",
            f"could not open Excel file: {ex.__class__.__name__}: {ex}",
        ))
        return result

    sheet_names = wb.sheetnames

    if "Individual" not in sheet_names and "Organisation" not in sheet_names:
        result.errors.append(Error(
            "(workbook)", 0, "-", "(sheets)", str(sheet_names),
            "workbook must contain at least one of 'Individual' or 'Organisation' sheets",
        ))
        return result

    if "Individual" in sheet_names:
        ws = wb["Individual"]
        _validate_headers(ws, "Individual", INDIVIDUAL_HEADERS, result)
        _parse_individual_sheet(ws, result)

    if "Organisation" in sheet_names:
        ws = wb["Organisation"]
        _validate_headers(ws, "Organisation", ORGANISATION_HEADERS, result)
        _parse_organisation_sheet(ws, result)

    return result